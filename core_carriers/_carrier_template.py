# Databricks notebook source
# KAP Data Platform — Core Carriers
# Shared helper functions for Bronze carrier ingestion
#
# Each carrier notebook does:
#   %run ../_carrier_template
# then calls one of the helpers below inside its @dp.table function.
#
# This keeps every carrier pipeline under 30 lines — adding carrier #6
# through #80 is just a new folder + a thin notebook that picks the
# right loader (csv / json / excel).

# COMMAND ----------

# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

from pyspark.sql import functions as F
from pyspark.sql.types import ArrayType, StringType

# COMMAND ----------

# ─── CATALOG / SCHEMA (carriers share one schema) ───

CATALOG = spark.conf.get("catalog", "kap_demo")
SCHEMA = "core_carriers"
LANDING_BASE = f"/Volumes/{CATALOG}/{SCHEMA}/carrier_file_landing"
SCHEMA_LOC_BASE = spark.conf.get("schema_location_base") + "/carriers"

# COMMAND ----------

def carrier_landing_path(carrier_folder):
    """Build the volume path for a specific carrier's files."""
    return f"{LANDING_BASE}/{carrier_folder}"


def carrier_schema_loc(carrier_folder):
    """Build the Auto Loader schema-tracking path for a carrier."""
    return f"{SCHEMA_LOC_BASE}/{carrier_folder}"

# COMMAND ----------

# ─── CSV LOADER ───

def load_csv_to_bronze(carrier_folder):
    """Read CSVs via Auto Loader and return a 5-column Bronze DataFrame.

    Good for: Acme, Liberty, Patriot, and most carriers who send flat files.
    """
    path = carrier_landing_path(carrier_folder)
    schema_loc = carrier_schema_loc(carrier_folder)

    raw = (
        spark.readStream
        .format("cloudFiles")
        .option("cloudFiles.format", "csv")
        .option("header", "true")
        .option("cloudFiles.inferColumnTypes", "false")
        .option("cloudFiles.schemaLocation", schema_loc)
        .load(path)
    )

    sorted_cols = sorted(raw.columns)
    raw_json = F.to_json(F.struct(*[F.col(c) for c in sorted_cols]))

    return (
        raw
        .withColumn("raw_row_variant", raw_json)
        .withColumn("file_path", F.col("_metadata.file_path"))
        .withColumn("ingestion_id", F.xxhash64(F.col("raw_row_variant"), F.col("file_path")))
        .withColumn("extracted_at", F.current_timestamp())
        .withColumn("source_row_hash", F.sha2(F.col("raw_row_variant"), 256))
        .select("ingestion_id", "file_path", "extracted_at", "source_row_hash", "raw_row_variant")
    )

# COMMAND ----------

# ─── JSON LOADER ───

def load_json_to_bronze(carrier_folder):
    """Read JSON-lines files via Auto Loader and return the Bronze schema.

    Good for: Summit and any carrier that sends a JSON feed.
    """
    path = carrier_landing_path(carrier_folder)
    schema_loc = carrier_schema_loc(carrier_folder)

    raw = (
        spark.readStream
        .format("cloudFiles")
        .option("cloudFiles.format", "json")
        .option("cloudFiles.inferColumnTypes", "false")
        .option("cloudFiles.schemaLocation", schema_loc)
        .load(path)
    )

    sorted_cols = sorted(raw.columns)
    raw_json = F.to_json(F.struct(*[F.col(c) for c in sorted_cols]))

    return (
        raw
        .withColumn("raw_row_variant", raw_json)
        .withColumn("file_path", F.col("_metadata.file_path"))
        .withColumn("ingestion_id", F.xxhash64(F.col("raw_row_variant"), F.col("file_path")))
        .withColumn("extracted_at", F.current_timestamp())
        .withColumn("source_row_hash", F.sha2(F.col("raw_row_variant"), 256))
        .select("ingestion_id", "file_path", "extracted_at", "source_row_hash", "raw_row_variant")
    )

# COMMAND ----------

# ─── EXCEL LOADER ───
# Same idea as sharepoint — crack the binary with openpyxl, dump rows as JSON.

@F.udf(returnType=ArrayType(StringType()))
def _parse_excel_rows(content):
    """Parse raw Excel bytes into a list of JSON strings (one per row)."""
    import io, json
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    result = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(row)]
            continue
        record = {h: str(v) if v is not None else None for h, v in zip(headers, row)}
        result.append(json.dumps(record, sort_keys=True))
    wb.close()
    return result


def load_excel_to_bronze(carrier_folder):
    """Read Excel files via binaryFile Auto Loader and return the Bronze schema.

    Good for: Beacon and any carrier that sends .xlsx bordereaux or schedules.
    """
    path = carrier_landing_path(carrier_folder)
    schema_loc = carrier_schema_loc(carrier_folder)

    raw = (
        spark.readStream
        .format("cloudFiles")
        .option("cloudFiles.format", "binaryFile")
        .option("cloudFiles.schemaLocation", schema_loc)
        .load(path)
    )

    parsed = (
        raw
        .withColumn("rows", _parse_excel_rows(F.col("content")))
        .select(
            F.col("path").alias("file_path"),
            F.explode("rows").alias("raw_row_variant"),
        )
    )

    return (
        parsed
        .withColumn("ingestion_id", F.xxhash64(F.col("raw_row_variant"), F.col("file_path")))
        .withColumn("extracted_at", F.current_timestamp())
        .withColumn("source_row_hash", F.sha2(F.col("raw_row_variant"), 256))
        .select("ingestion_id", "file_path", "extracted_at", "source_row_hash", "raw_row_variant")
    )
