"""Generate synthetic SharePoint files for the new_business domain.

Produces Excel files mimicking loss runs and submissions that agencies
upload to SharePoint folders. Each file has a unique schema — the Bronze
layer absorbs all differences into raw_row_variant.
"""

import os
from datetime import datetime, timedelta

from openpyxl import Workbook
from faker import Faker

from constants import (
    AGENCY_CODES, BUSINESS_NAMES, CARRIER_NAMES, COVERAGE_TYPES,
    US_STATES, seeded_random,
)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output", "sharepoint")
fake = Faker()
Faker.seed(42)
rng = seeded_random()


def generate_loss_runs():
    """Generate 3 loss-run Excel files, each from a different carrier format."""
    carriers = rng.sample(CARRIER_NAMES, 3)

    for i, carrier in enumerate(carriers, 1):
        wb = Workbook()
        ws = wb.active
        ws.title = "Loss Run"

        carrier_tag = carrier.split()[0].lower()
        filename = f"loss_run_{carrier_tag}_{datetime.now().strftime('%Y%m')}.xlsx"

        if i == 1:
            # Format A: detailed with claim numbers
            headers = [
                "Carrier", "Agency Code", "Insured Name", "Policy Number",
                "Claim Number", "Date of Loss", "Claimant", "Status",
                "Paid Amount", "Reserved Amount", "Coverage Type", "State",
            ]
            ws.append(headers)
            for _ in range(rng.randint(80, 150)):
                ws.append([
                    carrier,
                    rng.choice(AGENCY_CODES),
                    rng.choice(BUSINESS_NAMES),
                    f"POL-{rng.randint(100000, 999999)}",
                    f"CLM-{rng.randint(1000000, 9999999)}",
                    fake.date_between(start_date="-3y", end_date="today").isoformat(),
                    fake.name(),
                    rng.choice(["Open", "Closed", "Reopened"]),
                    round(rng.uniform(500, 250000), 2),
                    round(rng.uniform(0, 100000), 2),
                    rng.choice(COVERAGE_TYPES),
                    rng.choice(US_STATES),
                ])

        elif i == 2:
            # Format B: summary-level (no individual claims)
            headers = [
                "Carrier Name", "Producing Agent", "Named Insured",
                "Policy #", "Eff Date", "Exp Date", "Line of Business",
                "Total Incurred", "Loss Count", "Experience Mod",
            ]
            ws.append(headers)
            for _ in range(rng.randint(60, 120)):
                eff = fake.date_between(start_date="-2y", end_date="-6m")
                ws.append([
                    carrier,
                    rng.choice(AGENCY_CODES),
                    rng.choice(BUSINESS_NAMES),
                    f"{rng.randint(10000, 99999)}-{rng.randint(10, 99)}",
                    eff.isoformat(),
                    (eff + timedelta(days=365)).isoformat(),
                    rng.choice(COVERAGE_TYPES),
                    round(rng.uniform(0, 500000), 2),
                    rng.randint(0, 25),
                    round(rng.uniform(0.60, 1.80), 2),
                ])

        else:
            # Format C: minimal flat file
            headers = [
                "AGCY", "INSURED", "POL_NO", "LOSS_DT",
                "PAID", "OUTSTANDING", "TOTAL_INCURRED",
            ]
            ws.append(headers)
            for _ in range(rng.randint(50, 100)):
                paid = round(rng.uniform(0, 150000), 2)
                outstanding = round(rng.uniform(0, 80000), 2)
                ws.append([
                    rng.choice(AGENCY_CODES),
                    rng.choice(BUSINESS_NAMES),
                    f"P{rng.randint(1000000, 9999999)}",
                    fake.date_between(start_date="-3y", end_date="today").isoformat(),
                    paid,
                    outstanding,
                    paid + outstanding,
                ])

        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        print(f"✓ {filename} ({ws.max_row - 1} rows)")


def generate_submissions():
    """Generate 2 submission Excel files with different column layouts."""
    for i in range(1, 3):
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"

        filename = f"submissions_batch_{i}_{datetime.now().strftime('%Y%m')}.xlsx"

        if i == 1:
            headers = [
                "Submission ID", "Agency", "Business Name", "Contact Name",
                "Contact Email", "Coverage Requested", "Effective Date",
                "Estimated Premium", "Carrier Submitted To", "Status",
                "SIC Code", "State", "Annual Revenue",
            ]
            ws.append(headers)
            for _ in range(rng.randint(40, 80)):
                ws.append([
                    f"SUB-{rng.randint(10000, 99999)}",
                    rng.choice(AGENCY_CODES),
                    rng.choice(BUSINESS_NAMES),
                    fake.name(),
                    fake.email(),
                    rng.choice(COVERAGE_TYPES),
                    fake.date_between(start_date="today", end_date="+6m").isoformat(),
                    round(rng.uniform(2000, 150000), 2),
                    rng.choice(CARRIER_NAMES),
                    rng.choice(["Pending", "Quoted", "Bound", "Declined", "Lost"]),
                    rng.randint(1000, 9999),
                    rng.choice(US_STATES),
                    round(rng.uniform(100000, 50000000), 2),
                ])
        else:
            headers = [
                "Sub#", "Agent Code", "Insured", "LOB",
                "Target Carrier", "Requested Eff", "Status", "Notes",
            ]
            ws.append(headers)
            for _ in range(rng.randint(30, 60)):
                ws.append([
                    f"S{rng.randint(100000, 999999)}",
                    rng.choice(AGENCY_CODES),
                    rng.choice(BUSINESS_NAMES),
                    rng.choice(COVERAGE_TYPES),
                    rng.choice(CARRIER_NAMES),
                    fake.date_between(start_date="today", end_date="+6m").isoformat(),
                    rng.choice(["New", "In Progress", "Submitted", "Declined"]),
                    rng.choice(["", "Rush", "Renewal comparison", "Client referral"]),
                ])

        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        print(f"✓ {filename} ({ws.max_row - 1} rows)")


def main():
    """Generate all SharePoint fixture files."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("── Generating SharePoint files ──")
    generate_loss_runs()
    generate_submissions()
    print(f"✓ SharePoint fixtures written to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
