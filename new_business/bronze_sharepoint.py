# Databricks notebook source
# KAP Data Platform — New Business Domain
# Bronze: SharePoint file ingestion (Excel + CSV)

# COMMAND ----------

from pyspark import pipelines as dp
from pyspark.sql import functions as F
from pyspark.sql.types import ArrayType, StringType

# COMMAND ----------

# ─── CONFIG ───

CATALOG = spark.conf.get("catalog", "kap_demo")
SCHEMA = "new_business"
LANDING_PATH = f"/Volumes/{CATALOG}/{SCHEMA}/sharepoint_landing"
SCHEMA_LOC = spark.conf.get("schema_location_base") + "/sharepoint"

# COMMAND ----------

# ─── EXCEL PARSER UDF ───
# SharePoint drops Excel files with wildly different column layouts per carrier.
# We crack each file open with openpyxl, read every row, and dump it as a JSON
# string so the Bronze table doesn't care about schema differences at all.

@F.udf(returnType=ArrayType(StringType()))
def parse_excel(content):
    """Turn raw Excel bytes into a list of JSON strings, one per data row."""
    import io, json
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows_out = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # first row is always the header
            headers = [str(c).strip() if c is not None else f"col_{j}" for j, c in enumerate(row)]
            continue
        record = {}
        for h, val in zip(headers, row):
            record[h] = str(val) if val is not None else None
        rows_out.append(json.dumps(record, sort_keys=True))

    wb.close()
    return rows_out

# COMMAND ----------

# ─── BRONZE TABLE ───

@dp.table(name=f"{CATALOG}.{SCHEMA}.bronze_sharepoint", cluster_by_auto=True)
def bronze_sharepoint():
    """Ingest SharePoint Excel files into the standard 5-column Bronze schema."""
    raw = (
        spark.readStream
        .format("cloudFiles")
        .option("cloudFiles.format", "binaryFile")
        .option("cloudFiles.schemaLocation", SCHEMA_LOC)
        .load(LANDING_PATH)
    )

    # parse each Excel file, explode into individual rows
    parsed = (
        raw
        .withColumn("rows", parse_excel(F.col("content")))
        .select(
            F.col("path").alias("file_path"),
            F.explode("rows").alias("raw_row_variant"),
        )
    )

    # stamp the remaining Bronze columns
    bronze = (
        parsed
        .withColumn("ingestion_id", F.monotonically_increasing_id())
        .withColumn("extracted_at", F.current_timestamp())
        .withColumn("source_row_hash", F.sha2(F.col("raw_row_variant"), 256))
        .select("ingestion_id", "file_path", "extracted_at", "source_row_hash", "raw_row_variant")
    )

    return bronze
