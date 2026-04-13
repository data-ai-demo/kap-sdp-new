# Databricks notebook source
# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

# KAP Data Platform — Core Carriers
# Bronze: Beacon Mutual (Excel — workers comp premium bordereau)

from pyspark import pipelines as dp
from pyspark.sql import functions as F
from pyspark.sql.types import ArrayType, StringType

# COMMAND ----------

CATALOG = spark.conf.get("catalog", "kap_demo")
LANDING_PATH = f"/Volumes/{CATALOG}/core_carriers/carrier_file_landing/beacon"
SCHEMA_LOC = spark.conf.get("schema_location_base") + "/carriers/beacon"

# COMMAND ----------

@F.udf(returnType=ArrayType(StringType()))
def parse_excel(content):
    """Parse raw Excel bytes into a list of JSON strings (one per row)."""
    import io, json
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    result = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(row)]
            continue
        record = {h: str(v) if v is not None else None for h, v in zip(headers, row)}
        result.append(json.dumps(record, sort_keys=True))
    wb.close()
    return result

# COMMAND ----------

@dp.table(name="bronze_beacon", cluster_by_auto=True)
def bronze_beacon():
    """Beacon sends a monthly .xlsx bordereau with WC class codes,
    payroll, experience mod, schedule credits, and net premium.

    Excel files go through the binaryFile Auto Loader path — openpyxl
    cracks the workbook open and we explode each row into its own record.
    """
    raw = (
        spark.readStream
        .format("cloudFiles")
        .option("cloudFiles.format", "binaryFile")
        .option("cloudFiles.schemaLocation", SCHEMA_LOC)
        .load(LANDING_PATH)
    )

    parsed = (
        raw
        .withColumn("rows", parse_excel(F.col("content")))
        .select(
            F.col("path").alias("file_path"),
            F.explode("rows").alias("raw_row_variant"),
        )
    )

    return (
        parsed
        .withColumn("ingestion_id", F.xxhash64(F.col("raw_row_variant"), F.col("file_path")))
        .withColumn("extracted_at", F.current_timestamp())
        .withColumn("source_row_hash", F.sha2(F.col("raw_row_variant"), 256))
        .select("ingestion_id", "file_path", "extracted_at", "source_row_hash", "raw_row_variant")
    )
